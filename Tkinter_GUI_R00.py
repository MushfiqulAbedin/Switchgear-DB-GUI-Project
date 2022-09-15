from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
#from PIL import Image, ImageTk
#from xlrd import open_workbook
#import win32com.client
import openpyxl
from pathlib import Path

#import pandas as pd

# get the root window running
root = Tk()
root.title("Switchgear Database")
root.geometry("950x600")

# plugging the logo
'''
#img = PhotoImage(file="E:/Mushfiq/Sabbir Bhai Work/Database Project/frontend/logo.png")
#Label(root, image=img).grid(row=1, column=1)
width = 250
height = 150
img2 = Image.open("logo.jpg")
img2 = img2.resize((width, height), Image.ANTIALIAS)
photoImg = ImageTk.PhotoImage(img2)
Label(root, image=photoImg).grid(row=1, column=10)
'''

# Window for "Create New"
"""def newWindow():
    newWindow = Toplevel(root)
    newWindow.geometry("200x200")
    Label(newWindow, text="This is a new Window").pack(pady=10)
    newWindow.grab_set()                    # prevents interaction with the main window
"""

# Defining variables

var = StringVar()

def categoryWindow():
    categoryWindow = Toplevel(root)
    categoryWindow.geometry("350x200")
    categoryWindow.title("Category")
    categoryWindow.grab_set()
    Label(categoryWindow, text="MCCB Category").pack()

    # Trial method to check if the values are taken from the combobox
    def ratingWindow():
        print("method called")
        #print(catChosen.get())
        global var
        var = catChosen.get()
        print(var)
        print(type(var))
        # Rating Window
        ratingWindow = Toplevel(categoryWindow)
        ratingWindow.geometry("400x650")
        ratingWindow.title("Rating Details")
        ratingWindow.grab_set()

        # Level 2 Combobox - for rating selection
        Label(ratingWindow, text="Rating").pack()
        rat = tk.StringVar()
        ratChosen = ttk.Combobox(ratingWindow, width=10, textvariable=rat)
        ratChosen.bind("<<ComboboxSelected>>")
        ratChosen['values'] = ('800 A', '630 A', '400 A', '250 A', '160 A', '63 A', '32 A')
        ratChosen['state'] = 'readonly'
        ratChosen.pack()
        ratChosen.current()
        #Label(ratingWindow, text=str(rat.get())).pack()

        Label(ratingWindow, text="Pole").pack()
        pole = tk.StringVar()
        poleChosen = ttk.Combobox(ratingWindow, width=10, textvariable=pole)
        poleChosen.bind("<<ComboboxSelected>>")
        poleChosen['values'] = ('1P', '3P', '5P')
        poleChosen['state'] = 'readonly'
        poleChosen.pack()
        poleChosen.current()

        Label(ratingWindow, text="Breaking Capacity").pack()
        breaker = tk.StringVar()
        breakerChosen = ttk.Combobox(ratingWindow, width=10, textvariable=breaker)
        breakerChosen.bind("<<ComboboxSelected>>")
        breakerChosen['values'] = ('50 kA', '36 kA', '25 kA', '18 kA')
        breakerChosen['state'] = 'readonly'
        breakerChosen.pack()
        breakerChosen.current()

        # Method for making an input cell in a new window if a new tech description is to be added
        """def new(event):
            var = t1Chosen.get()
            print(var)
            if var == 'new...':
                descWindow = Toplevel(ratingWindow)
                descWindow.geometry("400x135")
                descWindow.title("New Technical Description")
                descWindow.grab_set()

                x = tk.StringVar()
                Entry(descWindow, textvariable=x).pack()
                Button(descWindow, text="Select").pack()"""

        def new1(event):
            print(t1Chosen.get())

        # Button command for New Technical Description
        def newSelect():
            t1Chosen['values'] += (newEntry.get(),)

        def newCode():
            temp = tk.StringVar()
            nCode = tk.StringVar()
            nCode = codeEntry.get()

            # Here, we check whether the code exists in database
            codefile = Path('./', 'Database.xlsx')
            wb = openpyxl.load_workbook(codefile)
            sht1 = wb.active

            # colname = 'Code Column'
            for row in sht1.values:                 # sht1.iter_rows(min_row=1, max_row=10000):
                for value in row:
                    if temp == "nok":
                        break
                    print(f"the row is {row}")
                    print(value)
                    print(nCode)
                    if value == nCode:
                        temp = "nok"
                    else:
                        temp = "ok"

            """xl = win32com.client.Dispatch("Excel.Application")
            codeFile = r'./Database.xlsx'
            wb = xl.Workbooks.Open(Filename=codeFile)
            sht1 = wb.Sheets.Item(1)
            match = sht1.Cells.Find(nCode)
            temp = match.Value
            print(temp)"""

            """book = open_workbook('Database.xlsx')
            book = pd.read_excel("Database.xlsx", engine=openpyxl)
            sheet = book.sheet_by_index(0)
            col = 'Code Column'

            for row in range(sheet.nrows):
                if sheet.cell(row, col).value == nCode:
                    temp = "ok"
                else:
                    temp = "nok" """

            if temp == "ok":
                Label(ratingWindow, text=f"Code {nCode} unused before", bg="green").pack()

            else:
                Label(ratingWindow, text="Code used before", bg="red").pack()

        Label(ratingWindow, text="Technical Description 1").pack()
        t1 = tk.StringVar()
        t1Chosen = ttk.Combobox(ratingWindow, width=20, textvariable=t1)
        t1Chosen.bind("<<ComboboxSelected>>", new1)
        t1Chosen['values'] = ('El Trip Unit PR121-LS/I', 'El Trip Unit PR122-LS/I', 'El Trip Unit PR123-LS/I',
                              'El Trip Unit PR221-LS/I')
        t1Chosen['state'] = 'readonly'      # Does not allow to manually set the value
        t1Chosen.pack()
        t1Chosen.current()

        newEntry =Entry(ratingWindow, bg='Yellow', width=15)
        newEntry.pack()

        Button(ratingWindow, text="Select", command=newSelect).pack()   # using button for getting the new tech description

        """if t1Chosen.get == 'new...':
            t1Chosen['state'] = 'normal'   # Allows to manually set the value
            #t1Chosen.set()
            newvar = tk.StringVar()
            t1Chosen['values'] = (*t1Chosen['values'], t1Chosen.get())
            print(t1Chosen['values'])"""

        Label(ratingWindow, text="Technical Description 2").pack()
        t2 = tk.StringVar()
        t2Chosen = ttk.Combobox(ratingWindow, width=10, textvariable=t2)
        t2Chosen.bind("<<ComboboxSelected>>")
        t2Chosen['values'] = ('Fixed')
        t2Chosen['state'] = 'readonly'
        t2Chosen.pack()
        t2Chosen.current()

        Label(ratingWindow, text="Code").pack()
        """code = tk.StringVar()
        codeChosen = ttk.Combobox(ratingWindow, width=10, textvariable=code)
        codeChosen.bind("<<ComboboxSelected>>")
        # codeChosen['values'] = ('Fixed')
        codeChosen['state'] = 'normal'
        codeChosen.pack()
        codeChosen.current()"""
        codeEntry = Entry(ratingWindow, bg='Yellow', width=15)
        codeEntry.pack()

        Button(ratingWindow, text="Select", command=newCode).pack()    # using button for adding code

        def saveRating():
            ratingVar = tk.StringVar()
            ratingVar = [catChosen.get(), ratChosen.get(), poleChosen.get(), breakerChosen.get(),
                         t1Chosen.get(), t2Chosen.get(), codeEntry.get()]
            r = ' '.join([str(elem) for elem in ratingVar])
            messagebox.showinfo(title="Device", message=f"The device saved is: {r}")



            #for row in sht1.rows:



        Button(ratingWindow, text="Save Rating", command=saveRating).pack()


    # Level 1 Combobox - for category selection
    cat = tk.StringVar()
    catChosen = ttk.Combobox(categoryWindow, width=15, textvariable=cat)
    catChosen.bind("<<ComboboxSelected>>")           # pass the method justamethod here to get the value
    catChosen['values'] = ('MCCB > 250 A', 'MCCB < 250 A')
    catChosen['state'] = 'readonly'
    catChosen.pack()
    catChosen.current()
    Label(categoryWindow, text=str(cat.get())).pack()

    """ Dummy method for selecting value from the combobox -- later done using justamethod
    def justamethod2():      
        print("something printed")
        print(var)
    """

    Button(categoryWindow, text="Select", command=ratingWindow).pack()



# Working in the Root Window
def new():
    Button(root, text="Create New Material", command=categoryWindow).grid(row=1, column=2)
    Button(root, text="Create New Project", command=categoryWindow).grid(row=2, column=2)

Button(root, text="New", command=new).grid(row=1, column=1)


#btn = Button(root, text="Open window", command=newWindow)
#btn.grid(row=50, column=50)


# infinite mainloop
root.mainloop()